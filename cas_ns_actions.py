"""
cas_ns_actions.py — Chargement de la table de correspondance "Motif NS -> Actions correctives"
(fichier cas_NS_colonnes_exactes_v3.xlsx), utilisée pour compléter automatiquement la colonne
"Actions correctives menées suite à l'audit..." des opérations Non satisfaisant.

Le fichier source doit se trouver dans le même dossier que ce module (déployé avec l'app).
"""

import re
from pathlib import Path

import openpyxl
import streamlit as st

from cee_lots_data import extract_fiche_code

NS_FILE_DEFAULT = str(Path(__file__).parent / "cas_NS_colonnes_exactes_v3.xlsx")

HEADER_ROW = 4       # ligne d'en-têtes du fichier de correspondance
FIRST_DATA_ROW = 5   # première ligne de données

COL_MOTIF = 1     # A : Motif NS (nom exact de la colonne dans la matrice)
COL_TRIGGER = 2   # B : Information déclenchant le NS
COL_LETTRE = 3    # C : Colonne (lettre, informative seulement)
COL_FICHES = 4    # D : Fiche CEE
COL_ACTION = 5    # E : A remplir dans la colonne


def _normalize_header(s):
    """Même logique que Traitement_Lot.normalize() : aplati les retours à la ligne et les
    espaces multiples, retire les astérisques de notes en tête."""
    if s is None:
        return ""
    s = str(s).replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s.lstrip("*").strip()


@st.cache_data(show_spinner=False)
def charger_mapping_ns(path: str = NS_FILE_DEFAULT):
    """Charge le fichier de correspondance en une liste de règles :
    [{"motif": <en-tête normalisé>, "trigger": <valeur déclenchante, en minuscules>,
      "fiches": [<codes fiche propres>], "action": <texte à insérer>}, ...]
    Retourne une liste vide (avec un avertissement Streamlit) si le fichier est introuvable."""
    if not Path(path).exists():
        st.warning(
            f"Fichier de correspondance des motifs NS introuvable ({Path(path).name}) — "
            "la colonne « Actions correctives » ne sera pas complétée automatiquement."
        )
        return []

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    mapping = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        motif = ws.cell(row=r, column=COL_MOTIF).value
        trigger = ws.cell(row=r, column=COL_TRIGGER).value
        fiche_cell = ws.cell(row=r, column=COL_FICHES).value
        action = ws.cell(row=r, column=COL_ACTION).value
        if not motif or not fiche_cell or not action:
            continue
        fiches = [extract_fiche_code(f) for f in str(fiche_cell).split(",") if f.strip()]
        fiches = [f for f in fiches if f]
        if not fiches:
            continue
        mapping.append(
            {
                "motif": _normalize_header(motif),
                "trigger": str(trigger).strip().lower() if trigger is not None else "",
                "fiches": fiches,
                "action": str(action).strip(),
            }
        )
    return mapping


def regles_pour_fiche(mapping, fiche_brute):
    """Filtre les règles de mapping applicables à une fiche donnée (code brut, éventuellement
    avec suffixe de variante — nettoyé via extract_fiche_code)."""
    code = extract_fiche_code(fiche_brute) if fiche_brute else None
    if not code:
        return []
    return [regle for regle in mapping if code in regle["fiches"]]


def action_corrective_pour_ligne(regles_fiche, valeurs_colonnes):
    """
    regles_fiche : règles de mapping déjà filtrées pour la fiche de la ligne (regles_pour_fiche).
    valeurs_colonnes : dict {en-tête normalisé de colonne: valeur brute de la cellule} pour les
        colonnes de la section "Données remplies par le bureau de contrôle...".
    Retourne le texte à écrire dans "Actions correctives..." (plusieurs motifs concaténés,
    un par ligne), ou None si aucun motif ne correspond.
    """
    textes = []
    for regle in regles_fiche:
        valeur_cellule = valeurs_colonnes.get(regle["motif"])
        if valeur_cellule is None:
            continue
        valeur_normalisee = str(valeur_cellule).strip().lower()
        if valeur_normalisee == regle["trigger"]:
            if regle["action"] not in textes:
                textes.append(regle["action"])
    return "\n".join(textes) if textes else None
