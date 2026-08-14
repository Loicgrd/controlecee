"""
Application Streamlit — Suivi et traitement des lots de contrôle CEE (fichier "Synthèse")

Étapes (après import du fichier "Synthèse - <LOT>.xlsx") :
1. Tableau de copie rapide pour Odicée (blocs "Contrôle sur site" / "Contrôle par contact").
2. Calcul de la "Surface retenue dans la demande" pour les fiches BAR-EN-101/102/103/105, et
   correction proportionnelle des volumes (kWh cumac) quand la surface retenue diffère de la
   surface déclarée.
3. Mise à jour de la/les fiche(s) BAR (export "Import lots de travaux") si une surface a diminué.
4. Taux de contrôle du lot, comparaison aux seuils réglementaires (arrêté du 27/07/2026) et
   catégorisation du lot (Cas 1/2/3).
5. Export Excel par bailleur, avec commentaire généré selon le cas du lot.
6. Génération des mails clients et ouverture d'un brouillon Outlook par bailleur.

Tableau Odicée : REFERENCE interne de l'opération + Conclusion de l'audit + Conclusion du
contrôle par contact.
"""

import base64
import io
import os
import re
import urllib.parse
from datetime import date
from decimal import ROUND_CEILING, ROUND_HALF_UP, Decimal
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl.styles import Alignment, Font, PatternFill

from cas_ns_actions import charger_mapping_ns, concatener_champ, regles_declenchees, regles_pour_fiche
from cee_lots_data import extract_fiche_code, get_seuils_fiche, parse_date_fr

st.set_page_config(page_title="Traitement des lots CEE", layout="wide")

FICHES_CONCERNEES = ("BAR-EN-101", "BAR-EN-102", "BAR-EN-103", "BAR-EN-105")


def sanitize_filename(name):
    return re.sub(r'[\\/*?:"<>|]', "", str(name)).strip()


# --------------------------------------------------------------------------------------
# Utilitaires génériques
# --------------------------------------------------------------------------------------

def normalize(s):
    """Nettoie un texte d'en-tête pour permettre des comparaisons robustes (retours à la
    ligne, astérisques de notes, espaces multiples...)."""
    if s is None:
        return ""
    s = str(s).replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s.lstrip("*").strip()


def find_header_row(ws, marker, max_row=10):
    """Repère la ligne d'en-têtes en cherchant une cellule contenant `marker`."""
    marker_n = normalize(marker).lower()
    for r in range(1, max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and marker_n in normalize(v).lower():
                return r
    return None


def build_header_map(ws, header_row):
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v:
            headers[c] = normalize(v)
    return headers


def _loosen(s):
    """Variante encore plus tolérante utilisée uniquement pour le rapprochement des
    en-têtes de colonnes : les tirets et espaces sont traités de façon équivalente
    (ex : 'non-qualité' == 'non qualité')."""
    return re.sub(r"[\s\-]+", " ", s).strip().lower()


def find_col(headers, prefix):
    """Retourne l'indice de la 1ère colonne dont l'en-tête commence par `prefix`."""
    prefix_n = _loosen(normalize(prefix))
    for c, h in headers.items():
        if _loosen(h).startswith(prefix_n):
            return c
    return None


def find_col_last(headers, prefix):
    """Retourne l'indice de la DERNIÈRE colonne dont l'en-tête commence par `prefix`.
    Utile quand un intitulé de colonne existe deux fois dans le fichier (ex : « Commentaires
    généraux » apparaît une fois sous « Données remplies par le bureau de contrôle » et une
    seconde fois, plus loin, sous « Données complétées par le demandeur »)."""
    prefix_n = _loosen(normalize(prefix))
    last = None
    for c, h in headers.items():
        if _loosen(h).startswith(prefix_n):
            last = c
    return last


def find_section_bounds(ws, section_prefix, header_row=1, max_col=None):
    """Retourne (col_debut, col_fin) de la section dont le titre (ligne `header_row`,
    généralement la ligne 1 des synthèses) commence par `section_prefix`. col_fin est la
    colonne juste avant le titre de section suivant, ou la dernière colonne du fichier s'il
    n'y en a pas. Retourne (None, None) si la section n'est pas trouvée."""
    max_col = max_col or ws.max_column
    section_starts = []
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if v and str(v).strip():
            section_starts.append(c)
    prefix_n = _loosen(normalize(section_prefix))
    for i, c in enumerate(section_starts):
        v = normalize(ws.cell(row=header_row, column=c).value)
        if _loosen(v).startswith(prefix_n):
            col_fin = section_starts[i + 1] - 1 if i + 1 < len(section_starts) else max_col
            return c, col_fin
    return None, None


def to_number(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    if s == "" or s.lower() in ("non vérifiable", "nv", "sans objet"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def fmt_like(original_value, new_value):
    """Formate `new_value` en respectant le type de la cellule d'origine (texte vs nombre),
    sans conserver de zéros inutiles (ex : 104.3 et non 104.30)."""
    new_value = round(new_value, 2)
    if isinstance(original_value, str):
        return f"{new_value:g}"
    return new_value


def round_half_up(x):
    """Arrondi à l'entier le plus proche, avec .5 arrondi vers le haut (contrairement à
    round() de Python qui arrondit .5 vers l'entier pair le plus proche)."""
    return int(Decimal(str(x)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def round_ceiling(x):
    """Arrondi systématiquement à l'entier supérieur."""
    return int(Decimal(str(x)).quantize(Decimal("1"), rounding=ROUND_CEILING))


# --------------------------------------------------------------------------------------
# UI
# --------------------------------------------------------------------------------------

st.title("Suivi des contrôles CEE — Synthèse QV580M")

synth_file = st.file_uploader(
    "1. Charger le fichier Synthèse (ex : Synthèse - QV580M.xlsx)", type=["xlsx"]
)

if not synth_file:
    st.info("Charge le fichier Synthèse pour commencer.")
    st.stop()

synth_bytes = synth_file.getvalue()

# Deux lectures : une pour les valeurs figées (calculs), une pour l'écriture (préserve les
# formules/format s'il y en a).
wb_write = openpyxl.load_workbook(io.BytesIO(synth_bytes), data_only=False)
wb_read = openpyxl.load_workbook(io.BytesIO(synth_bytes), data_only=True)
ws_write = wb_write.active
ws_read = wb_read.active

header_row = find_header_row(ws_read, "REFERENCE interne de l'opération")
if not header_row:
    st.error("Colonne \"REFERENCE interne de l'opération\" introuvable dans ce fichier.")
    st.stop()

headers = build_header_map(ws_read, header_row)

col_ref = find_col(headers, "REFERENCE interne de l'opération")
col_conclusion = find_col(headers, "Conclusion de l'audit")
col_conclusion_contact = find_col(headers, "Conclusion du contrôle par contact")
col_commentaires_generaux = find_col(headers, "Commentaires généraux")
col_non_qualite = find_col(headers, "Commentaire sur le type de non qualité relevée")
col_fiche = find_col(headers, "REFERENCE DE LA FICHE")
col_decl = find_col(headers, "Surface déclarée dans l'AH/facture")
col_mesuree = find_col(headers, "Surface mesurée par le bureau de contrôle")
col_estimee = find_col(headers, "Surface estimée par le bureau de contrôle")
col_retenue = find_col(headers, "Surface retenue dans la demande")
col_ecart = find_col(headers, "Écart entre surface mesurée")
col_vol_hp = find_col(headers, "VOLUME HORS PRECARITE")
col_vol_prec = find_col(headers, "VOLUME PRECARITE")


def cell_or_none(row, col):
    """Lit une cellule seulement si la colonne existe dans ce fichier ; None sinon.
    Utile pour les colonnes optionnelles (toutes les fiches n'ont pas de contrôle par
    contact, ni de colonnes de surface — certaines mesurent une longueur à la place)."""
    return ws_read.cell(row=row, column=col).value if col else None


# Certaines fiches n'ont pas de contrôle par contact (100% sur site), et certaines fiches
# (réseaux isolés, mesurées en longueur) n'ont pas du tout de colonnes de surface : ces
# groupes de colonnes sont donc optionnels, pas bloquants pour le reste de l'application.
has_contact = col_conclusion_contact is not None
has_surface_cols = all([col_decl, col_mesuree, col_estimee, col_retenue, col_ecart, col_vol_hp, col_vol_prec])

required = {
    "REFERENCE interne de l'opération": col_ref,
    "Conclusion de l'audit": col_conclusion,
    "REFERENCE DE LA FICHE": col_fiche,
}
missing = [name for name, c in required.items() if c is None]
if missing:
    st.error("Colonnes introuvables dans le fichier : " + ", ".join(missing))
    st.stop()

if not has_contact:
    st.info("Ce fichier n'a pas de colonne « Conclusion du contrôle par contact » — fiche à contrôle uniquement sur site.")
if not has_surface_cols:
    st.info(
        "Ce fichier n'a pas les colonnes de surface (déclarée/mesurée/estimée/retenue) — "
        "les étapes 2 et 3 (surface retenue, volumes, mise à jour fiche BAR) sont masquées "
        "(fiche non basée sur une surface, par ex. réseau isolé mesuré en longueur)."
    )

# Lignes dont la conclusion du contrôle sur site est renseignée (utilisées aussi pour le
# calcul de surface à l'étape 2, où seul le contrôle sur site mesure une surface).
rows = []
for r in range(header_row + 1, ws_read.max_row + 1):
    ref_val = ws_read.cell(row=r, column=col_ref).value
    concl_val = ws_read.cell(row=r, column=col_conclusion).value
    if ref_val is None or str(ref_val).strip() == "":
        continue
    if concl_val is None or str(concl_val).strip() == "":
        continue
    rows.append(r)

# Lignes dont la conclusion du contrôle par contact est renseignée (affichage uniquement).
rows_contact = []
if has_contact:
    for r in range(header_row + 1, ws_read.max_row + 1):
        ref_val = ws_read.cell(row=r, column=col_ref).value
        concl_val = ws_read.cell(row=r, column=col_conclusion_contact).value
        if ref_val is None or str(ref_val).strip() == "":
            continue
        if concl_val is None or str(concl_val).strip() == "":
            continue
        rows_contact.append(r)

st.success(
    f"{len(rows)} ligne(s) avec conclusion du contrôle sur site renseignée"
    + (f", {len(rows_contact)} ligne(s) avec conclusion du contrôle par contact renseignée." if has_contact else " (pas de contrôle par contact dans ce fichier).")
)

# --------------------------------------------------------------------------------------
# Étape 1 — Tableau à copier dans Odicée
# --------------------------------------------------------------------------------------

st.header(
    "1. Tableau à copier-coller dans Odicée",
    help=(
        "Deux blocs, un par type de contrôle, pour un copier-coller séparé :\n"
        "- Contrôle sur site : « REFERENCE interne » et « Conclusion du contrôle sur site » renseignées\n"
        "- Contrôle par contact : « REFERENCE interne » et « Conclusion du contrôle par contact » renseignées"
    ),
)

st.subheader("Bloc — Contrôle sur site")
odicee_site_data = [
    {
        "REFERENCE interne de l'opération": ws_read.cell(row=r, column=col_ref).value,
        "Conclusion du contrôle sur site": ws_read.cell(row=r, column=col_conclusion).value,
        "Commentaires généraux": cell_or_none(r, col_commentaires_generaux),
    }
    for r in rows
]
st.dataframe(
    pd.DataFrame(odicee_site_data),
    use_container_width=True,
    hide_index=True,
    column_config={
        "REFERENCE interne de l'opération": st.column_config.Column(
            help="Critère de tri : ligne affichée seulement si cette cellule ET « Conclusion du contrôle sur site » sont non vides."
        ),
        "Conclusion du contrôle sur site": st.column_config.Column(
            help="Colonne source dans le fichier : « Conclusion de l'audit ». C'est sa présence qui décide si la ligne apparaît dans ce bloc."
        ),
        "Commentaires généraux": st.column_config.Column(
            help="Affichée telle quelle (vide si non renseignée)."
        ),
    },
)

if has_contact:
    st.subheader("Bloc — Contrôle par contact")
    odicee_contact_data = [
        {
            "REFERENCE interne de l'opération": ws_read.cell(row=r, column=col_ref).value,
            "Conclusion du contrôle par contact": ws_read.cell(row=r, column=col_conclusion_contact).value,
            "Commentaire sur le type de non qualité relevée": cell_or_none(r, col_non_qualite),
        }
        for r in rows_contact
    ]
    st.dataframe(
        pd.DataFrame(odicee_contact_data),
        use_container_width=True,
        hide_index=True,
        column_config={
            "REFERENCE interne de l'opération": st.column_config.Column(
                help="Critère de tri : ligne affichée seulement si cette cellule ET « Conclusion du contrôle par contact » sont non vides."
            ),
            "Conclusion du contrôle par contact": st.column_config.Column(
                help="C'est la présence d'une valeur ici qui décide si la ligne apparaît dans ce bloc."
            ),
            "Commentaire sur le type de non qualité relevée": st.column_config.Column(
                help="Renseigné en général quand la conclusion est « Non satisfaisant » ; affiché tel quel (vide si non renseigné)."
            ),
        },
    )

if has_surface_cols:
    # --------------------------------------------------------------------------------------
    # Étape 2 — Surface retenue + volumes
    # --------------------------------------------------------------------------------------

    st.header(
        "2. Surface retenue dans la demande & correction des volumes",
        help=(
            "Parmi les lignes ci-dessus, seules celles dont « REFERENCE DE LA FICHE » commence par "
            "BAR-EN-101, BAR-EN-102, BAR-EN-103 ou BAR-EN-105 sont traitées.\n\n"
            "Règle de calcul de la surface retenue :\n"
            "1. Si Surface mesurée < Surface déclarée → retenue = Surface mesurée\n"
            "2. Sinon, si Écart > 10 % → retenue = Surface estimée\n"
            "3. Sinon → retenue = Surface déclarée\n\n"
            "Si la surface retenue diffère de la surface déclarée, les volumes sont recalculés "
            "au même prorata : VOLUME HORS PRECARITE est arrondi à l'entier supérieur, puis "
            "VOLUME PRECARITE = total recalculé (arrondi normalement) − VOLUME HORS PRECARITE."
        ),
    )
    st.caption(
        "Appliqué uniquement aux lignes BAR-EN-101 / BAR-EN-102 / BAR-EN-103 / BAR-EN-105 "
        "parmi les lignes ci-dessus."
    )

    changes = []
    skipped = []
    for r in rows:
        fiche_val = normalize(ws_read.cell(row=r, column=col_fiche).value or "")
        if not any(fiche_val.startswith(f) for f in FICHES_CONCERNEES):
            continue

        declaree = to_number(ws_read.cell(row=r, column=col_decl).value)
        mesuree = to_number(ws_read.cell(row=r, column=col_mesuree).value)
        estimee = to_number(ws_read.cell(row=r, column=col_estimee).value)
        ecart = to_number(ws_read.cell(row=r, column=col_ecart).value)
        ref_val = ws_read.cell(row=r, column=col_ref).value

        if declaree is None:
            skipped.append({"REFERENCE interne": ref_val, "Fiche": fiche_val, "Motif": "surface déclarée absente/non exploitable"})
            continue

        # Par défaut : surface déclarée. Uniquement remplacée si mesurée < déclarée, ou si
        # l'écart est strictement supérieur à 10 % (auquel cas on prend l'estimée).
        if mesuree is not None and mesuree < declaree:
            retenue = mesuree
        elif ecart is not None and ecart > 10 and estimee is not None:
            retenue = estimee
        else:
            retenue = declaree

        # Écriture de la surface retenue
        ws_write.cell(row=r, column=col_retenue).value = retenue

        decrease = retenue < declaree
        if retenue != declaree:
            ratio = retenue / declaree
            cell_hp = ws_write.cell(row=r, column=col_vol_hp)
            cell_prec = ws_write.cell(row=r, column=col_vol_prec)
            old_hp = to_number(ws_read.cell(row=r, column=col_vol_hp).value)
            old_prec = to_number(ws_read.cell(row=r, column=col_vol_prec).value)
            if old_hp is not None and old_prec is not None:
                # VOLUME HORS PRECARITE arrondi à l'entier supérieur, VOLUME PRECARITE déduit
                # du total (arrondi normalement) moins la valeur HORS PRECARITE déjà arrondie,
                # pour que la somme des deux reste cohérente avec le total recalculé.
                new_hp = round_ceiling(old_hp * ratio)
                new_total = round_half_up((old_hp + old_prec) * ratio)
                new_prec = new_total - new_hp
                cell_hp.value = new_hp
                cell_prec.value = new_prec
            elif old_hp is not None:
                cell_hp.value = round_half_up(old_hp * ratio)
            elif old_prec is not None:
                cell_prec.value = round_half_up(old_prec * ratio)

        id_lot = str(ref_val).rsplit("-", 1)[-1].strip()
        dossier = str(ref_val).split("-", 1)[0].strip()
        changes.append(
            {
                "REFERENCE interne": ref_val,
                "Dossier": dossier,
                "ID lot de travaux": id_lot,
                "Fiche": ws_read.cell(row=r, column=col_fiche).value,
                "Surface déclarée": declaree,
                "Surface mesurée": mesuree,
                "Surface estimée": estimee,
                "Écart (%)": ecart,
                "Surface retenue": retenue,
                "Diminution": "Oui" if decrease else "Non",
            }
        )

    any_decrease = any(c["Diminution"] == "Oui" for c in changes)

    if changes:
        with st.expander(f"Voir le détail des {len(changes)} ligne(s) traitées", expanded=False):
            st.dataframe(
                pd.DataFrame(changes),
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Dossier": st.column_config.Column(
                        help="Partie avant le « - » de REFERENCE interne de l'opération : sert de NUMERODOSSIER pour le lien Odicée à l'étape 3."
                    ),
                    "Fiche": st.column_config.Column(
                        help="Seules les fiches BAR-EN-101, BAR-EN-102, BAR-EN-103 et BAR-EN-105 sont retenues pour ce calcul."
                    ),
                    "Écart (%)": st.column_config.Column(
                        help="= (Surface déclarée − Surface mesurée) / Surface mesurée × 100. Détermine si on utilise la surface estimée quand mesurée ≥ déclarée (écart > 10 %)."
                    ),
                    "Surface retenue": st.column_config.Column(
                        help="Mesurée si mesurée < déclarée ; sinon estimée si écart > 10 % ; sinon déclarée."
                    ),
                    "Diminution": st.column_config.Column(
                        help="« Oui » si Surface retenue < Surface déclarée. Dès qu'au moins une ligne est à « Oui », l'étape 3 (import de fiche BAR) apparaît."
                    ),
                },
            )
    else:
        st.info("Aucune ligne BAR-EN-101/102/103/105 exploitable parmi les lignes filtrées.")

    if skipped:
        with st.expander(f"{len(skipped)} ligne(s) ignorée(s) (surface non exploitable)"):
            st.dataframe(pd.DataFrame(skipped), use_container_width=True, hide_index=True)

    # --------------------------------------------------------------------------------------
    # Étape 3 — Mise à jour de la/les fiche(s) BAR
    # --------------------------------------------------------------------------------------

    if any_decrease:
        st.header(
            "3. Mise à jour de la/les fiche(s) BAR",
            help=(
                "N'apparaît que si au moins une ligne a « Diminution » = Oui à l'étape 2.\n\n"
                "La correspondance entre la Synthèse et la fiche BAR se fait sur :\n"
                "REFERENCE interne de l'opération (partie après le dernier « - ») "
                "= ID lot de travaux de la fiche BAR."
            ),
        )
        st.caption(
            "Au moins une surface retenue est inférieure à la surface déclarée : "
            "importe la ou les fiches BAR (export \"Import lots de travaux\") pour y reporter "
            "la surface retenue, sur la ligne dont l'\"ID lot de travaux\" correspond."
        )

        surfaces_modifiees = [c for c in changes if c["Surface retenue"] != c["Surface déclarée"]]
        rows_to_apply = [
            c for c in surfaces_modifiees
            if c["Écart (%)"] is not None and c["Écart (%)"] <= 10
        ]

        dossier_fiches = {}
        for c in surfaces_modifiees:
            if c["Dossier"]:
                dossier_fiches.setdefault(c["Dossier"], set()).add(str(c["Fiche"]).strip())

        if dossier_fiches:
            st.markdown(
                "**Accès direct au(x) dossier(s) concerné(s) par une surface modifiée :**  \n"
                + "  \n".join(
                    f"- [{numero} ({', '.join(sorted(fiches))})](https://odicee.edf.fr/dossiers/{numero})"
                    for numero, fiches in sorted(dossier_fiches.items())
                )
            )

        bar_files = st.file_uploader(
            "Importer la ou les fiches BAR (ex : T155142_BAR-EN-102_A14_1.xlsx)",
            type=["xlsx"],
            accept_multiple_files=True,
            key="bar_upload",
        )

        if bar_files:
            for bf in bar_files:
                bar_bytes = bf.getvalue()
                bwb = openpyxl.load_workbook(io.BytesIO(bar_bytes), data_only=False)
                bws = bwb.active

                bar_header_row = find_header_row(bws, "ID lot de travaux")
                if not bar_header_row:
                    st.error(f"{bf.name} : colonne \"ID lot de travaux\" introuvable.")
                    continue

                bar_headers = build_header_map(bws, bar_header_row)
                col_id_lot = find_col(bar_headers, "ID lot de travaux")
                col_surface = find_col(bar_headers, "surface")

                if not col_id_lot or not col_surface:
                    st.error(f"{bf.name} : colonnes \"ID lot de travaux\" / \"surface\" introuvables.")
                    continue

                applied = []
                for r in range(bar_header_row + 1, bws.max_row + 1):
                    id_val = bws.cell(row=r, column=col_id_lot).value
                    id_num = to_number(id_val)
                    if id_num is None:
                        continue
                    id_lot_bar = str(int(id_num))

                    for chg in rows_to_apply:
                        if chg["ID lot de travaux"] == id_lot_bar:
                            cell = bws.cell(row=r, column=col_surface)
                            old_val = cell.value
                            new_val = fmt_like(old_val, chg["Surface retenue"])
                            cell.value = new_val
                            applied.append(
                                {
                                    "ID lot de travaux": id_lot_bar,
                                    "Ancienne surface": old_val,
                                    "Nouvelle surface": new_val,
                                }
                            )

                if applied:
                    st.write(f"**{bf.name}** — {len(applied)} ligne(s) mise(s) à jour :")
                    st.dataframe(pd.DataFrame(applied), use_container_width=True, hide_index=True)
                    out = io.BytesIO()
                    bwb.save(out)
                    st.download_button(
                        f"⬇️ Télécharger {bf.name} corrigé",
                        data=out.getvalue(),
                        file_name=f"MAJ_{bf.name}",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_{bf.name}",
                    )
                else:
                    st.info(f"{bf.name} : aucune ligne correspondant à un \"ID lot de travaux\" à corriger n'a été trouvée.")
else:
    changes = []
    any_decrease = False

# ========================================================================================
# Étape 4 — Taux de contrôle, conformité réglementaire et catégorisation du lot
# ========================================================================================

st.header(
    "4. Taux de contrôle et conformité du lot",
    help=(
        "Taux de satisfaisant (site ou contact) = nombre de satisfaisant / nombre total "
        "d'opérations du lot (toutes les lignes avec une REFERENCE interne).\n\n"
        "Taux de non satisfaisant sur site = nombre de non satisfaisant sur site / nombre "
        "de lignes avec une conclusion de contrôle sur site renseignée.\n\n"
        "Les seuils réglementaires sont recherchés dans la table de l'arrêté du 27/07/2026 "
        "pour la fiche BAR du lot, à la date d'engagement la plus récente."
    ),
)

# Toutes les opérations du lot (une ligne = une opération dès que REFERENCE interne est renseignée)
all_op_rows = []
for r in range(header_row + 1, ws_read.max_row + 1):
    ref_val = ws_read.cell(row=r, column=col_ref).value
    if ref_val is not None and str(ref_val).strip() != "":
        all_op_rows.append(r)

total_ops = len(all_op_rows)


def classify_conclusion(value):
    v = normalize(value).lower() if value else ""
    if not v:
        return "non_visite"
    if v == "satisfaisant":
        return "satisfaisant"
    if v == "non satisfaisant":
        return "non_satisfaisant"
    if "vérifiable" in v or "inaccessible" in v:
        return "inaccessible"
    return "autre"


nb_satisfaisant_site = 0
nb_non_satisfaisant_site = 0
nb_controles_site = 0
nb_satisfaisant_contact = 0
fiche_lot = None
dates_engagement = []
dates_achevement = []
cls_site_par_ligne = {}
actions_correctives_par_ligne = {}
actions_client_par_ligne = {}
motif_non_conformite_par_ligne = {}
mot_cle_par_ligne = {}

col_date_engagement = find_col(headers, "DATE D'ENGAGEMENT")
col_date_achevement = find_col(headers, "DATE d'achèvement de l'opération")
if not col_date_achevement and headers.get(18):
    # Repli positionnel : cette colonne se trouve historiquement en colonne R (18), même
    # quand son intitulé change d'un modèle de synthèse à l'autre (ex : "DATE DE LA FACTURE").
    col_date_achevement = 18

col_actions_correctives = find_col(headers, "Actions correctives menées suite à l'audit")
col_preciser = find_col(headers, "Préciser selon le cas si nécessaire")
col_commentaires_demandeur = find_col_last(headers, "Commentaires généraux")
has_ns_cols = all([col_actions_correctives, col_preciser, col_commentaires_demandeur])
if not has_ns_cols:
    st.info(
        "Colonnes « Actions correctives » / « Préciser selon le cas » / « Commentaires "
        "généraux (demandeur) » introuvables dans ce fichier — le remplissage automatique "
        "des motifs de non-satisfaisant est désactivé pour ce lot."
    )

bureau_controle_debut, bureau_controle_fin = find_section_bounds(
    ws_read, "Données remplies par le bureau de contrôle"
)
ns_mapping = charger_mapping_ns()

for r in all_op_rows:
    cls_site = classify_conclusion(ws_read.cell(row=r, column=col_conclusion).value)
    cls_site_par_ligne[r] = cls_site
    if cls_site != "non_visite":
        nb_controles_site += 1
    if cls_site == "satisfaisant":
        nb_satisfaisant_site += 1
    elif cls_site == "non_satisfaisant":
        nb_non_satisfaisant_site += 1

    if classify_conclusion(cell_or_none(r, col_conclusion_contact)) == "satisfaisant":
        nb_satisfaisant_contact += 1

    row_fiche_val = ws_read.cell(row=r, column=col_fiche).value
    if fiche_lot is None and row_fiche_val and str(row_fiche_val).strip():
        fiche_lot = str(row_fiche_val).strip()

    if col_date_engagement:
        d = parse_date_fr(ws_read.cell(row=r, column=col_date_engagement).value)
        if d:
            dates_engagement.append(d)

    if col_date_achevement:
        d2 = parse_date_fr(ws_read.cell(row=r, column=col_date_achevement).value)
        if d2:
            dates_achevement.append(d2)

    # Remplissage automatique de la colonne "Actions correctives" pour les opérations
    # Non satisfaisant sur site, à partir des colonnes de la section "Données remplies par
    # le bureau de contrôle..." et de la table de correspondance des motifs NS. On calcule
    # en une passe les 4 champs (texte synthèse, texte client, motif, mot clé) pour éviter
    # de refaire le matching plusieurs fois.
    if has_ns_cols and cls_site == "non_satisfaisant" and bureau_controle_debut and ns_mapping:
        regles_fiche = regles_pour_fiche(ns_mapping, row_fiche_val)
        if regles_fiche:
            valeurs_colonnes = {}
            for c in range(bureau_controle_debut, bureau_controle_fin + 1):
                h = headers.get(c)
                if h:
                    valeurs_colonnes[h] = ws_read.cell(row=r, column=c).value
            declenchees = regles_declenchees(regles_fiche, valeurs_colonnes)
            texte_synthese = concatener_champ(declenchees, "action_synthese")
            texte_client = concatener_champ(declenchees, "action_client")
            texte_motif = concatener_champ(declenchees, "motif_non_conformite")
            texte_mot_cle = concatener_champ(declenchees, "mot_cle")
            if texte_synthese:
                actions_correctives_par_ligne[r] = texte_synthese
                ws_write.cell(row=r, column=col_actions_correctives).value = texte_synthese
            if texte_client:
                actions_client_par_ligne[r] = texte_client
            if texte_motif:
                motif_non_conformite_par_ligne[r] = texte_motif
            if texte_mot_cle:
                mot_cle_par_ligne[r] = texte_mot_cle

# --------------------------------------------------------------------------------------
# Colonnes "Grand Précaire / Précaire / Classique" et "Version du coup de pouce (CDP...)",
# à remplir uniquement pour les fiches BAR-EN-101 et BAR-EN-103. Ajoutées à la suite de
# "Commentaires généraux" (section "Données complétées par le demandeur") si absentes.
# --------------------------------------------------------------------------------------
LIBELLE_GPE = (
    "Grand Précaire / Précaire / Classique à remplir différemment selon la période de la "
    "charte CDP (GPE/PE/CL pour les chartes 2018, 2019 et 2020, i.e. opérations engagées "
    "jusqu'au 31 mars 2021 ; GPE/CL pour les opérations engagées à partir du 1er avril 2021) "
    "Réaliser des tableaux séparés par période de charte"
)
LIBELLE_CDP = "Version du coup de pouce (CDP 2018 / CDP 2019 / CDP 2020 / CDP 2021) A remplir pour tout le tableau"

col_gpe = find_col(headers, "Grand Précaire / Précaire / Classique")
col_cdp = find_col(headers, "Version du coup de pouce")

if has_ns_cols and (not col_gpe or not col_cdp):
    if col_commentaires_demandeur == ws_write.max_column:
        insert_at = col_commentaires_demandeur + 1
        ws_write.insert_cols(insert_at, amount=2)
        col_gpe = col_gpe or insert_at
        col_cdp = col_cdp or (insert_at + 1)
        ws_write.cell(row=header_row, column=col_gpe).value = LIBELLE_GPE
        ws_write.cell(row=header_row, column=col_cdp).value = LIBELLE_CDP
    else:
        st.info(
            "Colonnes « Grand Précaire / Précaire / Classique » / « Version du coup de "
            "pouce » absentes et non ajoutées automatiquement (« Commentaires généraux » "
            "n'est pas la dernière colonne du fichier)."
        )
        col_gpe = col_cdp = None

if has_ns_cols and col_gpe and col_cdp:
    date_engagement_min = min(dates_engagement) if dates_engagement else None
    for r in all_op_rows:
        row_fiche_code = extract_fiche_code(ws_read.cell(row=r, column=col_fiche).value)
        if row_fiche_code not in ("BAR-EN-101", "BAR-EN-103"):
            continue
        if date_engagement_min and date_engagement_min <= date(2021, 3, 31):
            ws_write.cell(row=r, column=col_gpe).value = "GPE/PE/CL"
        else:
            ws_write.cell(row=r, column=col_gpe).value = "GPE/CL"
        ws_write.cell(row=r, column=col_cdp).value = "SO"

taux_s_site = (nb_satisfaisant_site / total_ops * 100) if total_ops else 0.0
taux_s_contact = (nb_satisfaisant_contact / total_ops * 100) if total_ops else 0.0
taux_ns_site = (nb_non_satisfaisant_site / nb_controles_site * 100) if nb_controles_site else 0.0
date_engagement_max = max(dates_engagement) if dates_engagement else None
date_achevement_min = min(dates_achevement) if dates_achevement else None

seuil_site, seuil_contact = get_seuils_fiche(fiche_lot, date_engagement_max) if fiche_lot and date_engagement_max else (None, None)
seuils_trouves = seuil_site is not None or seuil_contact is not None

# --- Conformité du taux de satisfaisant (nécessaire ici, avant la catégorisation Cas 1/2/3
# affichée plus bas, car elle détermine aussi le remplissage des lignes non visitées) ---
if not seuils_trouves:
    # Aucun seuil réglementaire trouvé pour cette fiche à cette date d'engagement (table de
    # l'arrêté non renseignée pour cette période, ex : engagement antérieur à 2022) : on ne
    # peut pas garantir la conformité, donc par prudence on considère que ce n'est PAS conforme
    # plutôt que l'inverse.
    site_ok = False
    contact_ok = False
else:
    site_ok = True if seuil_site is None else (taux_s_site >= seuil_site)
    if seuil_contact is None:
        contact_ok = True
    else:
        contact_ok_direct = taux_s_contact >= seuil_contact
        if seuil_site is not None:
            contact_ok = contact_ok_direct or ((taux_s_site + taux_s_contact) >= (seuil_site + seuil_contact))
        else:
            contact_ok = contact_ok_direct
taux_satisfaisant_conforme = site_ok and contact_ok

# Lignes non visitées (Conclusion de l'audit vide) OU non vérifiables : si un des taux de
# satisfaisant n'est pas atteint (donc dans le cas 3), on retire ces opérations du dossier —
# et on vide les colonnes qui ne doivent plus être renseignées pour une opération retirée.
col_raison_sociale_demandeur = find_col(headers, "RAISON SOCIALE du demandeur")
col_siren_demandeur = find_col(headers, "SIREN du demandeur")
colonnes_a_vider = [c for c in (col_raison_sociale_demandeur, col_siren_demandeur, col_vol_hp, col_vol_prec) if c]

preciser_par_ligne = {}
commentaires_demandeur_par_ligne = {}
if has_ns_cols and not taux_satisfaisant_conforme:
    for r in all_op_rows:
        if cls_site_par_ligne.get(r) in ("non_visite", "inaccessible"):
            preciser_par_ligne[r] = "Opération retirée du dossier"
            commentaires_demandeur_par_ligne[r] = "Un des taux règlementaires n'est pas atteint"
            ws_write.cell(row=r, column=col_preciser).value = preciser_par_ligne[r]
            ws_write.cell(row=r, column=col_commentaires_demandeur).value = commentaires_demandeur_par_ligne[r]
            for c in colonnes_a_vider:
                ws_write.cell(row=r, column=c).value = None

# Fichier Synthèse mis à jour (surfaces/volumes + motifs NS + lignes retirées), prêt à télécharger
buf_synth = io.BytesIO()
wb_write.save(buf_synth)
synth_path = Path(synth_file.name)
synth_v2_name = f"{synth_path.stem} V2{synth_path.suffix}"
st.download_button(
    "⬇️ Télécharger la Synthèse mise à jour",
    data=buf_synth.getvalue(),
    file_name=synth_v2_name,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

col_info1, col_info2, col_info3, col_info4 = st.columns(4)
with col_info1:
    st.metric("Fiche BAR du lot", fiche_lot or "—")
with col_info2:
    st.metric("Date d'engagement la plus récente", date_engagement_max.strftime("%d/%m/%Y") if date_engagement_max else "—")
with col_info3:
    st.metric(
        "Date de réalisation la plus ancienne",
        date_achevement_min.strftime("%d/%m/%Y") if date_achevement_min else "—",
        help="Date la plus ancienne de la colonne « DATE d'achèvement de l'opération ».",
    )
with col_info4:
    seuil_txt = ""
    if seuil_site is not None:
        seuil_txt += f"Site ≥ {seuil_site:g}%"
    if seuil_contact is not None:
        seuil_txt += (" · " if seuil_txt else "") + f"Contact ≥ {seuil_contact:g}%"
    st.metric("Seuils réglementaires trouvés", seuil_txt or "non trouvés")

if fiche_lot and date_engagement_max and seuil_site is None and seuil_contact is None:
    st.warning(
        f"Aucun seuil trouvé dans la table de l'arrêté pour la fiche « {fiche_lot} » à la date "
        f"{date_engagement_max.strftime('%d/%m/%Y')} (période antérieure à la table, ou fiche "
        "absente). Par prudence, le taux de satisfaisant est donc considéré comme **non "
        "conforme** tant qu'aucun seuil n'est confirmé. Vérifie le code de fiche ou complète la "
        "table dans cee_lots_data.py."
    )

seuil_ns_max = st.number_input(
    "Seuil maximal de non satisfaisant sur site (%)",
    min_value=0.0, max_value=100.0, value=14.0, step=0.5,
    help="Modifiable ponctuellement. Le taux de non satisfaisant du lot ne doit pas dépasser ce seuil.",
)

# --- Conformité du taux de non satisfaisant ---
ns_conforme = taux_ns_site <= seuil_ns_max

# --- Catégorisation du lot ---
if taux_satisfaisant_conforme and ns_conforme:
    cas_lot = 1
    conclusion_cas = "La totalité des opérations dans le lot est déposable."
elif taux_satisfaisant_conforme and not ns_conforme:
    cas_lot = 2
    conclusion_cas = "On peut déposer toutes les opérations visitées seulement."
else:
    cas_lot = 3
    conclusion_cas = "On ne peut déposer que les opérations contrôlées satisfaisantes et non satisfaisantes."

def carte_taux(titre, ok, valeur_txt, sous_texte):
    color = "#C6EFCE" if ok else "#FFCCCC"
    icone = "✅" if ok else "❌"
    return (
        f"<div style='background:{color};padding:12px;border-radius:8px;text-align:center'>"
        f"<b>{titre}</b> {icone}<br>"
        f"<span style='font-size:22px;font-weight:bold'>{valeur_txt}</span><br>"
        f"<small>{sous_texte}</small>"
        f"</div>"
    )


taux_cols = st.columns(3)
with taux_cols[0]:
    sous = (f"{nb_satisfaisant_site}/{total_ops} — seuil ≥ {seuil_site:g}%" if seuil_site is not None
            else f"{nb_satisfaisant_site}/{total_ops} — pas de seuil pour cette fiche")
    st.markdown(
        carte_taux("Taux satisfaisant sur site", site_ok, f"{taux_s_site:.1f} %", sous),
        unsafe_allow_html=True,
    )
with taux_cols[1]:
    sous = (f"{nb_satisfaisant_contact}/{total_ops} — seuil ≥ {seuil_contact:g}%" if seuil_contact is not None
            else f"{nb_satisfaisant_contact}/{total_ops} — pas de seuil pour cette fiche")
    st.markdown(
        carte_taux("Taux satisfaisant par contact", contact_ok, f"{taux_s_contact:.1f} %", sous),
        unsafe_allow_html=True,
    )
with taux_cols[2]:
    sous = f"{nb_non_satisfaisant_site}/{nb_controles_site} — seuil ≤ {seuil_ns_max:g}%"
    st.markdown(
        carte_taux("Taux non satisfaisant sur site", ns_conforme, f"{taux_ns_site:.1f} %", sous),
        unsafe_allow_html=True,
    )

cas_couleur = {1: "#e8f5e9", 2: "#fff3e0", 3: "#ffebee"}[cas_lot]
st.markdown(
    f"<div style='background-color:{cas_couleur}; padding:14px; border-radius:8px;'>"
    f"<b>Cas {cas_lot}</b> — {conclusion_cas}</div>",

    unsafe_allow_html=True,
)

# ========================================================================================
# Étape 5 — Export Excel et mail par bailleur
# ========================================================================================

st.header(
    "5. Export Excel et mail par bailleur",
    help=(
        "Un fichier par « RAISON SOCIALE du bénéficiaire de l'opération », avec les colonnes "
        "Numéro dossier ODICEE / Adresse / Code postal / Ville / Conclusion sur site / "
        "Conclusion par contact / Commentaire / Action(s) corrective(s) nécessaire(s) "
        "(généré selon le cas du lot).\n\n"
        "Le bouton mail ouvre le client mail par défaut (objet + corps pré-remplis) ; la pièce "
        "jointe n'est pas ajoutée automatiquement (limite du lien mailto) — télécharge-la et "
        "joins-la manuellement."
    ),
)

col_i = find_col(headers, "RAISON SOCIALE du bénéficiaire")
col_e = find_col(headers, "NOM DU SITE")
col_f = find_col(headers, "ADRESSE de l'opération")
col_g = find_col(headers, "CODE POSTAL")
col_h = find_col(headers, "VILLE")

if not all([col_i, col_e, col_f, col_g, col_h]):
    st.error("Colonnes bailleur/adresse introuvables (RAISON SOCIALE / NOM DU SITE / ADRESSE / CODE POSTAL / VILLE).")
else:
    def build_commentaire(cls_site, commentaires_generaux, cas):
        """Retourne (texte du commentaire, statut) où statut détermine la couleur de la
        ligne dans l'export. « Transféré » utilise le même texte et la même couleur, que
        l'opération soit inaccessible/non vérifiable ou non visitée."""
        if cls_site == "satisfaisant":
            return "L'opération sera valorisée dans ce lot", "valorisee"
        if cls_site == "non_satisfaisant":
            txt = str(commentaires_generaux).strip() if commentaires_generaux else ""
            return (txt or "Non satisfaisant sur site — voir Commentaires généraux"), "non_satisfaisant"
        if cls_site == "inaccessible":
            if cas in (1, 2):
                return "L'opération sera valorisée dans ce lot", "valorisee"
            return "L'opération sera transférée dans un nouveau lot de contrôle si la date de fin de validité du dossier nous le permet", "transferee"
        # non_visite
        if cas == 1:
            return "L'opération sera valorisée dans ce lot", "valorisee"
        return "L'opération sera transférée dans un nouveau lot de contrôle si la date de fin de validité du dossier nous le permet", "transferee"

    bailleurs = {}
    for r in all_op_rows:
        bailleur = ws_read.cell(row=r, column=col_i).value
        bailleur = str(bailleur).strip() if bailleur else "(bailleur non renseigné)"

        concl_site_val = ws_read.cell(row=r, column=col_conclusion).value
        concl_contact_val = cell_or_none(r, col_conclusion_contact)
        cls_site = classify_conclusion(concl_site_val)
        commentaires_generaux_val = cell_or_none(r, col_commentaires_generaux)
        commentaire_txt, statut = build_commentaire(cls_site, commentaires_generaux_val, cas_lot)
        dossier_num = str(ws_read.cell(row=r, column=col_ref).value or "").split("-", 1)[0].strip()

        bailleurs.setdefault(bailleur, []).append(
            {
                "Numéro dossier ODICEE": dossier_num,
                "Adresse": ws_read.cell(row=r, column=col_f).value,
                "Code postal": ws_read.cell(row=r, column=col_g).value,
                "Ville": ws_read.cell(row=r, column=col_h).value,
                "Conclusion du contrôle sur site": str(concl_site_val).strip() if concl_site_val and str(concl_site_val).strip() else "Non visité",
                "Conclusion du contrôle par contact": str(concl_contact_val).strip() if concl_contact_val and str(concl_contact_val).strip() else "Non visité",
                "Commentaire": commentaire_txt,
                "Action(s) corrective(s) nécessaire(s)": actions_client_par_ligne.get(r, ""),
                "Dossier": dossier_num,
                "Statut": statut,
            }
        )

    COULEUR_COMMENTAIRE = {
        "valorisee": "C6EFCE",
        "non_satisfaisant": "FFC7CE",
        "transferee": "FFE0B2",
    }

    def build_excel_bailleur(lignes):
        wb_b = openpyxl.Workbook()
        ws_b = wb_b.active
        ws_b.title = "Résultats contrôle"
        headers_b = list(lignes[0].keys())
        headers_b = [h for h in headers_b if h not in ("Dossier", "Statut")]  # colonnes techniques, pas affichées
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2F5496")
        for c, h in enumerate(headers_b, 1):
            cell = ws_b.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r, ligne in enumerate(lignes, 2):
            for c, h in enumerate(headers_b, 1):
                cell = ws_b.cell(row=r, column=c, value=ligne[h])
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            color = COULEUR_COMMENTAIRE.get(ligne["Statut"])
            if color:
                for c in range(1, len(headers_b) + 1):
                    ws_b.cell(row=r, column=c).fill = PatternFill("solid", fgColor=color)
        widths = [22, 30, 12, 18, 22, 22, 45, 45]
        for i, w in enumerate(widths[: len(headers_b)], 1):
            ws_b.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        buf_b = io.BytesIO()
        wb_b.save(buf_b)
        return buf_b.getvalue()

    st.session_state["bailleurs_data"] = bailleurs
    st.session_state["cas_lot"] = cas_lot
    st.session_state["conclusion_cas"] = conclusion_cas

    guess_lot = re.sub(r"^Synth[eè]se[\s\-_]+", "", Path(synth_file.name).stem, flags=re.IGNORECASE).strip()
    num_lot = st.text_input("Numéro de lot (pour l'objet du mail)", value=guess_lot)

    def construire_corps_mail(cas, num_lot, site_ok, ns_depasse, seuil_ns, cases, saisie_lot, saisie_dossier, operations_controlees, dossiers):
        lignes = ["Bonjour,", ""]
        lignes.append(f"Pour votre information, nous avons reçu le retour du lot de contrôle {num_lot}.")
        lignes.append("")
        lignes.append("Résultat des dossiers : Vous trouverez ci-joint les résultats des contrôles pour vos opérations.")
        lignes.append("")
        if dossiers:
            lignes.append("Liste des dossiers concernés :")
            for d in dossiers:
                lignes.append(f"- {d}")
            lignes.append("")
        lignes.append("Résultat du Lot :")
        lignes.append("")
        lignes.append(f"Taux de visite satisfaisante sur site : {'Atteint' if site_ok else 'Non-Atteint'}")
        if ns_depasse:
            lignes.append(f"Taux de visite non satisfaisante sur site supérieur aux {seuil_ns:g}% autorisés")
        lignes.append("")
        lignes.append("Conclusion :")
        lignes.append("")

        if cas == 1:
            lignes.append(
                "- Toutes les opérations (« Satisfaisant », « Non Vérifiables », « Non visité » "
                "et « Non Satisfaisant », mais mise en conformité avant la date du prochain "
                "dépôt au PNCEE) peuvent être validées dans ce lot."
            )
        elif cas == 2:
            lignes.append(
                "- Les opérations contrôlées (« Satisfaisant », « Non Vérifiables » et « Non "
                "Satisfaisant » mais mise en conformité avant la date du prochain dépôt au "
                "PNCEE) peuvent être validées dans ce lot."
            )
            lignes.append(
                "- Les opérations « non visitées » ne peuvent pas être validées dans ce lot, "
                f"car le taux d'opérations contrôlées non-satisfaisantes est supérieur à {seuil_ns:g} %."
            )
        else:
            lignes.append(
                "- Les opérations contrôlées (« Satisfaisant », et « Non Satisfaisant » mais "
                "mise en conformité avant la date du prochain dépôt au PNCEE) peuvent être "
                "validées dans ce lot."
            )
            lignes.append(
                "- Les opérations « non visitées » et « non vérifiable » ne peuvent pas être "
                "validées dans ce lot, car le taux d'opérations contrôlées non-satisfaisantes "
                f"est supérieur à {seuil_ns:g} %."
            )

        if cas in (2, 3):
            if cases.get("lot_destination"):
                lignes.append(
                    "- Les opérations non valorisées dans ce lot ont été transférées dans le "
                    f"nouveau lot de contrôle {saisie_lot}."
                )
            if cases.get("dossier_destination"):
                lignes.append(
                    "- Les opérations non valorisées dans ce lot ont été transférées dans le "
                    f"nouveau dossier {saisie_dossier}."
                )

        if cases.get("delai_insuffisant"):
            lignes.append(
                "- La date de fin de travaux du dossier inférieure à 3 mois et ne nous permet "
                "pas de lancer un nouveau contrôle."
            )
        if cases.get("ah_non_recue"):
            lignes.append("- L'attestation sur l'honneur n'a toujours pas été reçue.")
        if cases.get("document_non_conforme"):
            lignes.append("- Des documents sur le dossier ne sont pas conformes.")

        lignes.append("")
        if operations_controlees:
            lignes.append(
                "Les rapports de contrôles sont disponibles en téléchargement dans les pièces "
                "jointes de chaque dossier ODICEE."
            )
            lignes.append("")
        lignes.append("Votre interlocuteur EDF et nous-même restons disponibles.")
        lignes.append("")
        lignes.append("Bien à vous,")
        return "\n".join(lignes)

    def construire_corps_mail_html(cas, num_lot, site_ok, ns_depasse, seuil_ns, cases, saisie_lot, saisie_dossier, operations_controlees, dossiers):
        """Même contenu que construire_corps_mail, mais en HTML avec la mise en forme du
        modèle Word (titres bleus soulignés, ligne du taux en vert/rouge selon atteint ou
        non, conclusion en bleu)."""
        def titre(txt):
            return f'<p style="color:#4C94D8;text-decoration:underline;font-weight:bold;margin:18px 0 8px 0;line-height:1.5;">{txt}</p>'

        def para(txt):
            return f'<p style="margin:10px 0;line-height:1.5;">{txt}</p>'

        html = ['<div style="font-family:Calibri,Arial,sans-serif;font-size:11pt;color:#000000;line-height:1.5;">']
        html.append(para("Bonjour,"))
        html.append(para(f"Pour votre information, nous avons reçu le retour du lot de contrôle <b>{num_lot}</b>."))
        html.append(titre("Résultat des dossiers"))
        html.append(para("Vous trouverez ci-joint les résultats des contrôles pour vos opérations."))
        if dossiers:
            html.append(para("<b>Liste des dossiers concernés :</b>"))
            html.append('<ul style="margin:6px 0;padding-left:22px;">')
            for d in dossiers:
                html.append(f'<li style="margin:2px 0;line-height:1.4;">{d}</li>')
            html.append("</ul>")
        html.append(titre("Résultat du Lot"))
        if site_ok:
            html.append(f'<p style="color:#6FC040;font-weight:bold;margin:10px 0;line-height:1.5;">Taux de visite satisfaisante sur site : Atteint</p>')
        else:
            html.append(f'<p style="color:#C00000;font-weight:bold;margin:10px 0;line-height:1.5;">Taux de visite satisfaisante sur site : Non-Atteint</p>')
        if ns_depasse:
            html.append(f'<p style="color:#C82613;font-weight:bold;margin:10px 0;line-height:1.5;">Taux de visite non satisfaisante sur site supérieur aux {seuil_ns:g}% autorisés</p>')
        html.append(titre("Conclusion"))

        puces = []
        if cas == 1:
            puces.append(
                "Toutes les opérations (« Satisfaisant », « Non Vérifiables », « Non visité » "
                "et « Non Satisfaisant », mais mise en conformité avant la date du prochain "
                "dépôt au PNCEE) peuvent être validées dans ce lot."
            )
        elif cas == 2:
            puces.append(
                "Les opérations contrôlées (« Satisfaisant », « Non Vérifiables » et « Non "
                "Satisfaisant » mais mise en conformité avant la date du prochain dépôt au "
                "PNCEE) peuvent être validées dans ce lot."
            )
            puces.append(
                "Les opérations « non visitées » ne peuvent pas être validées dans ce lot, "
                f"car le taux d'opérations contrôlées non-satisfaisantes est supérieur à {seuil_ns:g} %."
            )
        else:
            puces.append(
                "Les opérations contrôlées (« Satisfaisant », et « Non Satisfaisant » mais "
                "mise en conformité avant la date du prochain dépôt au PNCEE) peuvent être "
                "validées dans ce lot."
            )
            puces.append(
                "Les opérations « non visitées » et « non vérifiable » ne peuvent pas être "
                "validées dans ce lot, car le taux d'opérations contrôlées non-satisfaisantes "
                f"est supérieur à {seuil_ns:g} %."
            )

        if cas in (2, 3):
            if cases.get("lot_destination"):
                puces.append(
                    "Les opérations non valorisées dans ce lot ont été transférées dans le "
                    f"nouveau lot de contrôle {saisie_lot}."
                )
            if cases.get("dossier_destination"):
                puces.append(
                    "Les opérations non valorisées dans ce lot ont été transférées dans le "
                    f"nouveau dossier {saisie_dossier}."
                )

        if cases.get("delai_insuffisant"):
            puces.append(
                "La date de fin de travaux du dossier inférieure à 3 mois et ne nous permet "
                "pas de lancer un nouveau contrôle."
            )
        if cases.get("ah_non_recue"):
            puces.append("L'attestation sur l'honneur n'a toujours pas été reçue.")
        if cases.get("document_non_conforme"):
            puces.append("Des documents sur le dossier ne sont pas conformes.")

        html.append('<ul style="margin:10px 0;padding-left:22px;color:#1F6FD6;">')
        for p in puces:
            html.append(f'<li style="margin:6px 0;line-height:1.5;">{p}</li>')
        html.append("</ul>")

        if operations_controlees:
            html.append(para(
                "Les rapports de contrôles sont disponibles en téléchargement dans les pièces "
                "jointes de chaque dossier ODICEE."
            ))
        html.append(para("Votre interlocuteur EDF et nous-même restons disponibles."))
        html.append(para("Bien à vous,"))
        html.append("</div>")
        return "".join(html)

    def construire_eml(destinataire, cc, sujet, corps_html, piece_jointe_bytes, piece_jointe_nom):
        msg = MIMEMultipart()
        msg["Subject"] = sujet
        if destinataire:
            msg["To"] = destinataire
        if cc:
            msg["Cc"] = cc
        msg.attach(MIMEText(corps_html, "html", "utf-8"))
        part = MIMEApplication(
            piece_jointe_bytes,
            _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            Name=piece_jointe_nom,
        )
        part["Content-Disposition"] = f'attachment; filename="{piece_jointe_nom}"'
        msg.attach(part)
        return msg.as_bytes()

    for bailleur, lignes_b in sorted(bailleurs.items()):
        cle = sanitize_filename(bailleur)
        st.markdown(f"**🏢 {bailleur}** — {len(lignes_b)} opération(s)")

        case_cols = st.columns(5) if cas_lot in (2, 3) else st.columns(3)
        cases = {}
        i = 0
        if cas_lot in (2, 3):
            with case_cols[i]:
                cases["lot_destination"] = st.checkbox("Lot de destination", key=f"cb_lot_dest_{cle}")
            i += 1
            with case_cols[i]:
                cases["dossier_destination"] = st.checkbox("Dossier de destination", key=f"cb_dossier_dest_{cle}")
            i += 1
        with case_cols[i]:
            cases["delai_insuffisant"] = st.checkbox("Délais insuffisant", key=f"cb_delai_{cle}")
        i += 1
        with case_cols[i]:
            cases["ah_non_recue"] = st.checkbox("Ah non reçue", key=f"cb_ah_{cle}")
        i += 1
        with case_cols[i]:
            cases["document_non_conforme"] = st.checkbox("Document non conforme", key=f"cb_doc_{cle}")

        saisie_lot = saisie_dossier = ""
        if cases.get("lot_destination"):
            saisie_lot = st.text_input("Numéro du nouveau lot de contrôle", key=f"saisie_lot_{cle}")
        if cases.get("dossier_destination"):
            saisie_dossier = st.text_input("Numéro du nouveau dossier", key=f"saisie_dossier_{cle}")

        operations_controlees = any(l["Conclusion du contrôle sur site"] != "Non visité" for l in lignes_b)
        dossiers_bailleur = sorted({l["Dossier"] for l in lignes_b if l.get("Dossier")})
        corps = construire_corps_mail(
            cas_lot, num_lot, site_ok, not ns_conforme, seuil_ns_max, cases, saisie_lot, saisie_dossier, operations_controlees, dossiers_bailleur
        )
        corps_html = construire_corps_mail_html(
            cas_lot, num_lot, site_ok, not ns_conforme, seuil_ns_max, cases, saisie_lot, saisie_dossier, operations_controlees, dossiers_bailleur
        )
        subject = f"Retour de contrôle {num_lot}"
        attach_name = f"{sanitize_filename(num_lot)} - {cle}.xlsx"
        xlsx_bytes_b = build_excel_bailleur(lignes_b)
        cc_address = "controle.ceebs@promotelec-services.com"
        mailto_url = (
            f"mailto:?cc={urllib.parse.quote(cc_address)}&subject={urllib.parse.quote(subject)}"
        )
        corps_html_b64 = base64.b64encode(corps_html.encode("utf-8")).decode("ascii")

        col_name, col_actions = st.columns([3, 2])
        with col_name:
            with st.expander("Aperçu du corps du mail"):
                st.text(corps)
        with col_actions:
            b64_xlsx = base64.b64encode(xlsx_bytes_b).decode("ascii")
            components.html(
                f"""
                <div style="display:flex; gap:12px; font-family:'Source Sans Pro', sans-serif;">
                  <a id="mail-btn" href="{mailto_url}" target="_blank" rel="noopener"
                     onclick="var b=document.getElementById('mail-btn'); b.style.background='#c6efce'; b.style.borderColor='#4caf50'; b.innerHTML='✅ Mail ouvert';"
                     style="flex:1; text-align:center; padding:0.55em 1em; border-radius:8px;
                            border:1px solid #d3d3d3; background:#f0f2f6; color:#31333F;
                            text-decoration:none; font-size:14px; cursor:pointer; transition:background 0.15s;">
                     📧 Envoyer le mail
                  </a>
                  <a id="dl-btn" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_xlsx}"
                     download="{attach_name}"
                     onclick="var b=document.getElementById('dl-btn'); b.style.background='#c6efce'; b.style.borderColor='#4caf50'; b.innerHTML='✅ Excel téléchargé';"
                     style="flex:1; text-align:center; padding:0.55em 1em; border-radius:8px;
                            border:1px solid #d3d3d3; background:#f0f2f6; color:#31333F;
                            text-decoration:none; font-size:14px; cursor:pointer; transition:background 0.15s;">
                     ⬇️ Télécharger l'Excel
                  </a>
                </div>
                """,
                height=60,
            )
            components.html(
                f"""
                <div style="display:flex; gap:12px; font-family:'Source Sans Pro', sans-serif; margin-top:6px; padding:2px;">
                  <button id="copy-html-btn"
                     style="width:100%; box-sizing:border-box; text-align:center; padding:0.6em 1em;
                            border-radius:8px; border:1px solid #d3d3d3; background:#f0f2f6;
                            color:#31333F; font-size:14px; line-height:1.4; cursor:pointer;
                            transition:background 0.15s; white-space:normal;">
                     📋 Copier le texte mis en forme
                  </button>
                </div>
                <script>
                  (function() {{
                    var b64 = "{corps_html_b64}";
                    function b64ToUtf8(b) {{
                      var binary = atob(b);
                      var bytes = new Uint8Array(binary.length);
                      for (var i = 0; i < binary.length; i++) {{ bytes[i] = binary.charCodeAt(i); }}
                      return new TextDecoder('utf-8').decode(bytes);
                    }}
                    var htmlContent = b64ToUtf8(b64);
                    var plainContent = htmlContent.replace(/<[^>]+>/g, '');
                    var btn = document.getElementById('copy-html-btn');

                    function ok() {{ btn.style.background='#c6efce'; btn.style.borderColor='#4caf50'; btn.innerHTML='✅ Copié — colle avec Ctrl+V dans le mail'; }}
                    function ko(detail) {{
                      btn.style.background='#ffcccc'; btn.style.borderColor='#c00000';
                      btn.innerHTML='❌ Copie impossible sur ce navigateur';
                      console.error('Copie du mail impossible :', detail);
                    }}

                    function copierViaSelection() {{
                      var container = document.createElement('div');
                      container.setAttribute('contenteditable', 'true');
                      container.style.position = 'fixed';
                      container.style.left = '-9999px';
                      container.style.top = '0';
                      container.innerHTML = htmlContent;
                      document.body.appendChild(container);
                      var selection = window.getSelection();
                      var range = document.createRange();
                      range.selectNodeContents(container);
                      selection.removeAllRanges();
                      selection.addRange(range);
                      var reussi = false;
                      try {{ reussi = document.execCommand('copy'); }} catch (err) {{ reussi = false; }}
                      selection.removeAllRanges();
                      document.body.removeChild(container);
                      return reussi;
                    }}

                    function copierViaClipboardApi() {{
                      try {{
                        var blobHtml = new Blob([htmlContent], {{type: 'text/html'}});
                        var blobText = new Blob([plainContent], {{type: 'text/plain'}});
                        return navigator.clipboard.write([new ClipboardItem({{'text/html': blobHtml, 'text/plain': blobText}})]);
                      }} catch (err) {{
                        return Promise.reject(err);
                      }}
                    }}

                    btn.addEventListener('click', function() {{
                      if (copierViaSelection()) {{
                        ok();
                        return;
                      }}
                      if (navigator.clipboard && window.ClipboardItem) {{
                        copierViaClipboardApi().then(ok, function(err) {{
                          if (navigator.clipboard.writeText) {{
                            navigator.clipboard.writeText(plainContent).then(ok, function(err2) {{ ko(err2); }});
                          }} else {{ ko(err); }}
                        }});
                      }} else {{
                        ko('execCommand et Clipboard API indisponibles');
                      }}
                    }});
                  }})();
                </script>
                """,
                height=90,
            )
        st.divider()

    # ========================================================================================
    # Étape 6 — Tableau NS (à copier-coller, sans en-têtes)
    # ========================================================================================

    st.header(
        "6. Tableau NS à copier-coller",
        help=(
            "Une ligne par opération « Non satisfaisant » (Conclusion de l'audit), avec les "
            "colonnes du modèle « Modèle_tableau_NS.xlsx ». Le bouton copie le tableau SANS "
            "en-têtes, prêt à coller dans un Excel existant."
        ),
    )

    def find_col_exact(headers, texte):
        """Comme find_col, mais exige une correspondance EXACTE (pas juste un préfixe) —
        utile pour un en-tête court comme « SIREN » qui serait sinon confondu avec « SIREN
        du demandeur », « SIREN du professionnel », etc."""
        cible = _loosen(normalize(texte))
        for c, h in headers.items():
            if _loosen(h) == cible:
                return c
        return None

    col_siren_beneficiaire = find_col_exact(headers, "SIREN")
    col_raison_professionnel = find_col(headers, "RAISON SOCIALE du professionnel")
    col_siret_professionnel = find_col(headers, "SIRET de l'entreprise ayant réalisé l'opération")

    colonnes_ns_manquantes = [
        nom for nom, c in [
            ("SIREN (bénéficiaire)", col_siren_beneficiaire),
            ("RAISON SOCIALE du professionnel", col_raison_professionnel),
            ("SIRET de l'entreprise ayant réalisé l'opération", col_siret_professionnel),
        ] if not c
    ]
    if colonnes_ns_manquantes:
        st.info("Colonnes introuvables pour le tableau NS : " + ", ".join(colonnes_ns_manquantes) + " — les cellules correspondantes resteront vides.")

    lignes_ns = []
    for r in all_op_rows:
        if cls_site_par_ligne.get(r) != "non_satisfaisant":
            continue
        vol_hp_val = to_number(ws_read.cell(row=r, column=col_vol_hp).value) if col_vol_hp else None
        vol_prec_val = to_number(ws_read.cell(row=r, column=col_vol_prec).value) if col_vol_prec else None
        volume_mwhc = ""
        if vol_hp_val is not None or vol_prec_val is not None:
            volume_mwhc = round(((vol_hp_val or 0) + (vol_prec_val or 0)) * 0.001, 3)

        fiche_brute_ligne = ws_read.cell(row=r, column=col_fiche).value
        fiche_code_ligne = extract_fiche_code(fiche_brute_ligne) or ""
        fiche_bar_sans_prefixe = re.sub(r"^(BAR|BAT)-", "", fiche_code_ligne)

        # Les cellules concaténées (plusieurs motifs) utilisent un retour à la ligne en
        # interne, ce qui casse le copier-coller dans Excel (crée des lignes en trop) : on
        # remplace par un point-virgule pour cette table spécifiquement.
        motif_txt = motif_non_conformite_par_ligne.get(r, "")
        mot_cle_txt = mot_cle_par_ligne.get(r, "")
        motif_pour_copie = motif_txt.replace("\n", "; ")
        mots_cle_liste = mot_cle_txt.split("\n") if mot_cle_txt else []
        mots_cle_liste = (mots_cle_liste + ["", "", "", ""])[:4]

        date_debut_travaux = ""
        if col_date_engagement:
            d_eng = parse_date_fr(ws_read.cell(row=r, column=col_date_engagement).value)
            date_debut_travaux = d_eng.strftime("%d/%m/%Y") if d_eng else ""

        date_realisation_travaux = ""
        if col_date_achevement:
            d_ach = parse_date_fr(ws_read.cell(row=r, column=col_date_achevement).value)
            date_realisation_travaux = d_ach.strftime("%d/%m/%Y") if d_ach else ""

        lignes_ns.append(
            [
                ws_read.cell(row=r, column=col_ref).value or "",
                ws_read.cell(row=r, column=col_f).value or "",
                ws_read.cell(row=r, column=col_g).value or "",
                ws_read.cell(row=r, column=col_h).value or "",
                ws_read.cell(row=r, column=col_i).value or "",
                (ws_read.cell(row=r, column=col_siren_beneficiaire).value or "") if col_siren_beneficiaire else "",
                (ws_read.cell(row=r, column=col_raison_professionnel).value or "") if col_raison_professionnel else "",
                (ws_read.cell(row=r, column=col_siret_professionnel).value or "") if col_siret_professionnel else "",
                volume_mwhc,
                "",  # Dossier d'origine de l'opération (Emmy)
                num_lot,  # Lot de contrôle d'origine de l'opération
                date.today().strftime("%d/%m/%Y"),  # Date de demande de mise en conformité
                motif_pour_copie,
                "en cours",  # Actions correctives mises en œuvre
                "",  # Type d'action corrective mise en œuvre
                "",  # Date de réalisation des actions correctives
                "",  # Référence interne de destination de l'opération
                "",  # Dossier de destination de l'opération corrigée (Emmy)
                "",  # Lot de contrôle secondaire de l'opération
                mot_cle_txt.replace("\n", "; "),
                "",  # Délai de réalisation des correctifs (en jours)
                fiche_bar_sans_prefixe,
                mots_cle_liste[0],
                mots_cle_liste[1],
                mots_cle_liste[2],
                mots_cle_liste[3],
                date_debut_travaux,
                date_realisation_travaux,
            ]
        )

    entetes_ns = [
        "Référence de l'opération", "Adresse de l'opération", "Code Postal", "Ville",
        "Beneficiaire de l'opération", "SIREN benéficiaire", "Professionnel titulaire du signe de qualité",
        "Siret Professionnel", "Volume en MWhc", "Dossier d'origine de l'opération (Emmy)",
        "Lot de contrôle d'origine de l'opération", "Date de demande de mise en conformité",
        "Motif de non-conformité", "Actions correctives mises en œuvre",
        "Type d'action corrective mise en œuvre", "Date de réalisation des actions correctives",
        "Référence interne de destination de l'opération", "Dossier de destination de l'opération corrigée (Emmy)",
        "Lot de contrôle secondaire de l'opération", "Mot clé", "Délai de réalisation des correctifs (en jours)",
        "Fiche BAR", "Mot clé 1", "Mot clé 2", "Mot clé 3", "Mot clé 4",
        "Date de début de travaux", "Date de réalisation des travaux",
    ]

    if lignes_ns:
        st.dataframe(pd.DataFrame(lignes_ns, columns=entetes_ns), use_container_width=True, hide_index=True)

        tsv = "\n".join("\t".join(str(v) for v in ligne) for ligne in lignes_ns)
        tsv_js = tsv.replace("\\", "\\\\").replace("`", "\\`").replace("${", "\\${")
        components.html(
            f"""
            <div style="font-family:'Source Sans Pro', sans-serif;">
              <button id="copy-ns-btn"
                 onclick="navigator.clipboard.writeText(`{tsv_js}`); var b=document.getElementById('copy-ns-btn'); b.style.background='#c6efce'; b.style.borderColor='#4caf50'; b.innerHTML='✅ Tableau copié';"
                 style="padding:0.55em 1.2em; border-radius:8px; border:1px solid #d3d3d3;
                        background:#f0f2f6; color:#31333F; font-size:14px; cursor:pointer;
                        transition:background 0.15s;">
                 📋 Copier le tableau (sans en-têtes)
              </button>
            </div>
            """,
            height=50,
        )
    else:
        st.info("Aucune opération « Non satisfaisant » dans ce lot — rien à copier.")

